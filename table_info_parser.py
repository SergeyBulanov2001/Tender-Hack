import random

import openpyxl
from openpyxl_image_loader import SheetImageLoader


class TableInfoParser:
    def __init__(self, file):
        self.file = file

    def get_header(self):
        wb = openpyxl.load_workbook(self.file)
        ws = wb.active

        cell = ws.cell(row=1, column=1)
        column_counter = 1

        final_array = []

        while cell.value is not None:
            final_array.append(cell.value)
            column_counter += 1
            cell = ws.cell(row=1, column=column_counter)

        return final_array

    def get_random_row(self):
        wb = openpyxl.load_workbook(self.file)
        ws = wb.active

        table_height = 2

        cell = ws.cell(row=table_height, column=1)

        while cell.value is not None:
            table_height += 1
            cell = ws.cell(row=table_height, column=1)

        cell = ws.cell(row=1, column=1)
        column_counter = 1

        rand_row = random.randint(2, table_height - 1)

        print(rand_row)

        final_array = {}

        while cell.value is not None:
            column_counter += 1
            final_array[cell.value] = ws.cell(row=rand_row, column=column_counter).value
            cell = ws.cell(row=1, column=column_counter)

        return final_array


