import openpyxl
from openpyxl_image_loader import SheetImageLoader
import random
import string


class Parser:
    def __init__(self, xlsx_file, xlsx_info):
        self.file = xlsx_file
        self.info = xlsx_info

    def parse(self):

        wb = openpyxl.load_workbook(self.file)
        ws = wb.active

        column_counter = 1
        row_counter = 2

        finished = False

        last_cell = ws.cell(column=len(self.info), row=1)

        if last_cell.value is None:
            raise Exception("Inconsistency of the fields in the table and specified on the site")

        while not finished:
            final_array = []
            params = {}

            cell_content = ws.cell(column=column_counter, row=row_counter)
            cell_type = self.info[column_counter - 1]["type"]

            if cell_content.value is None and cell_type != "image":
                finished = True
            else:
                while cell_content.value is not None or cell_type == "image":

                    if column_counter > len(self.info):
                        break

                    if not self.info[column_counter - 1]:
                        column_counter += 1
                        continue

                    cell_content = ws.cell(column=column_counter, row=row_counter)
                    cell_type = self.info[column_counter - 1]["type"]

                    if cell_type == "image":
                        image_loader = SheetImageLoader(ws)
                        image = image_loader.get(cell_content.coordinate)
                        image_token = self.generate_token(8)
                        image.save("data/images/{}.png".format(image_token))
                        cur_cell = self.info[column_counter - 1].copy()
                        cur_cell["content"] = "data/images/{}.png".format(image_token)
                        final_array.append(cur_cell)

                    elif cell_type == "param":
                        params = self.specifications(cell_content.value)

                    else:
                        cur_cell = self.info[column_counter - 1].copy()
                        cur_cell["content"] = cell_content.value

                        final_array.append(cur_cell)
                    column_counter += 1

            column_counter = 1
            row_counter += 1
            yield final_array, params

    def get_categories(self):

        category_column = None
        row_counter = 2

        for item in self.info:
            if not item: continue
            if item["type"] == "category":
                category_column = self.info.index(item) + 1

        if category_column is None:
            raise Exception('Category column not marked!')

        wb = openpyxl.load_workbook(self.file)
        ws = wb.active

        cell_content = ws.cell(column=category_column, row=row_counter)

        final_array = []

        while cell_content.value is not None:
            final_array.append(cell_content.value)
            row_counter += 1
            cell_content = ws.cell(column=category_column, row=row_counter)

        return set(final_array)

    def specifications(self, text, line_break=".\n", split_param=": "):
        params = {}
        for param in text.split(line_break):
            param = param.split(split_param)


            if len(param) == 2:
                try:

                    float(param[1].split()[0])
                    params[param[0]] = param[1].split()

                except:
                    params[param[0]] = param
            else:
                print("Error")

        return params

    def generate_token(self, length):
        alphabet = string.ascii_letters + string.digits
        token = ''
        for i in range(length):
            token += alphabet[random.randint(0, len(alphabet) - 1)]
        return token


if __name__ == '__main__':
    # TODO проверка одинаковых типов

    name = "name"
    company = "company"
    parser = Parser("data/esus.xlsx",
                    [{"type": "1", "unit": "1"}, {"type": "2", "unit": "1"}, {"type": "3", "unit": "1"},
                     {"type": "4", "unit": "1"}, {"type": "5", "unit": "1"}, {"type": "6", "unit": "1"},
                     None, {"type": "8", "unit": "1"}, {"type": "image", "unit": "1"},
                     {"type": "10", "unit": "1"}, {"type": "11", "unit": "1"}, {"type": "12", "unit": "1"},
                     {"type": "13", "unit": "1"}, {"type": "14", "unit": "1"}, {"type": "15", "unit": "1"},
                     {"type": "16", "unit": "1"
                      }, {"type": "category", "unit": "1"}])

    for i in parser.parse():
        print(i)

    print(1)
    print(parser.get_categories())
